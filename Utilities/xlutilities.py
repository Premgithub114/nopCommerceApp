import openpyxl

def rownum(path,sheet):
    book = openpyxl.load_workbook(path)
    sht = book.get_sheet_by_name(sheet)
    num = sht.max_row
    return (num)
def colnum(path,sheet):
    book=openpyxl.load_workbook(path)
    sht=book.get_sheet_by_name(sheet)
    num = sht.max_column
    return (num)
def read_data(path,sheet,ro,co):
    book = openpyxl.load_workbook(path)
    sht = book.get_sheet_by_name(sheet)
    return (sht.cell(row=ro,column=co).value)
def write_data(path,sheet,ro,co,data):
    book = openpyxl.load_workbook(path)
    sht = book.get_sheet_by_name(sheet)
    sht.cell(row=ro, column=co).value=data
    book.save(path)

